#!/usr/bin/env python3
# coding:UTF-8
import rclpy
from rclpy.node import Node
import serial
from sensor_msgs.msg import NavSatFix
from std_msgs.msg import Header
from geometry_msgs.msg import Point, TransformStamped
import math
from tf2_ros import TransformBroadcaster
from openpyxl import Workbook  # type: ignore
import os
# Para manipular archivos Excel

class GPSNode(Node):
    def __init__(self):
        super().__init__('gps_node')
        
        # Publicadores para datos GPS
        self.publisher_ = self.create_publisher(NavSatFix, 'gps/fix', 10)
        self.cartesian_publisher_ = self.create_publisher(Point, 'gps/cartesian', 10)

        # TF broadcaster para publicar las transformaciones entre `map` y `base_link`
        self.tf_broadcaster = TransformBroadcaster(self)
        
        # Parámetros del temporizador
        timer_period = 0.1  # Segundos entre publicaciones
        self.timer = self.create_timer(timer_period, self.GPS_read)
        
        # Configuración del puerto serie (GPS)
        port = '/dev/ttyUSB0'  # Asegúrate de que este sea el puerto correcto
        baud = 9600
        try:
            self.ser = serial.Serial(port, baud, timeout=2)
            self.get_logger().info(f"GPS Serial Opened! Baudrate={baud}")
        except serial.SerialException as e:
            self.get_logger().error(f"Failed to open GPS Serial: {e}")
            raise

        # Punto de referencia inicial
        self.lat0 = 0.0
        self.lon0 = 0.0
        self.R = 6371000  # Radio de la Tierra en metros

        # Configuración del archivo Excel
        #self.excel_file = "gps_data.xlsx"
        self.excel_file = "/home/gps_data.xlsx"

        self.init_excel_file()

    def init_excel_file(self):
        # Crear el archivo Excel y configurar las columnas
        if not os.path.exists(self.excel_file):
            self.workbook = Workbook()
            sheet = self.workbook.active
            sheet.title = "GPS Data"
            sheet.append(["Latitude", "Longitude", "Altitude"])  # Encabezados
            self.workbook.save(self.excel_file)
            self.get_logger().info(f"Excel file {self.excel_file} created.")
        else:
            self.workbook = Workbook()
            self.workbook = Workbook().load_workbook(self.excel_file)
            self.get_logger().info(f"Excel file {self.excel_file} loaded.")

    def GPS_read(self):
        try:
            if self.ser.in_waiting:
                line = self.ser.readline().decode('ascii', errors='replace').strip()

                # Filtra solo las líneas GGA
                if line.startswith('$GNGGA'):
                    self.get_logger().info(f"Raw GGA GPS Data: {line}")
                    GGA = line.split(',')

                    if len(GGA) < 15:
                        self.get_logger().warn("GPS GGA data incomplete")
                        return

                    # Extraer latitud, longitud y altitud
                    lat = self.convert_to_degrees(GGA[2], GGA[3], True)
                    lon = self.convert_to_degrees(GGA[4], GGA[5], False)
                    alt = float(GGA[9]) if GGA[9] else 0.0
                    self.publish_gps_data(lat, lon, alt)
        except Exception as e:
            self.get_logger().error(f"Error processing GPS data: {e}")

    def convert_to_degrees(self, raw_value, direction, is_latitude):
        if not raw_value:
            return 0.0

        if is_latitude:
            degrees = float(raw_value[:2])
            minutes = float(raw_value[2:]) / 60.0
        else:
            degrees = float(raw_value[:3])
            minutes = float(raw_value[3:]) / 60.0

        result = degrees + minutes
        if direction in ['S', 'W']:
            result *= -1
        return result

    def publish_gps_data(self, lat, lon, alt):
        # Publicar mensaje NavSatFix
        msg = NavSatFix()
        msg.header = Header()
        msg.header.stamp = self.get_clock().now().to_msg()
        msg.header.frame_id = 'map'  # Correcto: el GPS da posición absoluta
        
        msg.latitude = lat
        msg.longitude = lon
        msg.altitude = alt

        # Covarianza de la posición
        precision_meters = 1.0
        msg.position_covariance = [
            precision_meters**2, 0.0, 0.0,
            0.0, precision_meters**2, 0.0,
            0.0, 0.0, precision_meters**2
        ]
        msg.position_covariance_type = NavSatFix.COVARIANCE_TYPE_APPROXIMATED

        # Publicar datos GPS
        self.publisher_.publish(msg)

        # Conversión a coordenadas cartesianas
        x = self.R * (lat - self.lat0) * (math.pi / 180) * math.cos(self.lat0 * (math.pi / 180))
        y = self.R * (lon - self.lon0) * (math.pi / 180)

        # Publicar coordenadas cartesianas
        cartesian_msg = Point()
        cartesian_msg.x = x
        cartesian_msg.y = y
        cartesian_msg.z = alt  # Usa la altitud como coordenada z

        self.cartesian_publisher_.publish(cartesian_msg)
        # Escribir los datos en el archivo Excel
        self.write_to_excel(lat, lon, alt)

        self.get_logger().info(f"Published GPS data: lat={lat}, lon={lon}, alt={alt}")
        self.get_logger().info(f"Published Cartesian GPS data: x={x}, y={y}, z={alt}")

    def write_to_excel(self, lat, lon, alt):
        try:
            sheet = self.workbook.active
            sheet.append([lat, lon, alt])  # Agregar los datos
            self.workbook.save(self.excel_file)
            self.get_logger().info(f"GPS data written to Excel: lat={lat}, lon={lon}, alt={alt}")
        except Exception as e:
            self.get_logger().error(f"Error writing data to Excel: {e}")

def main(args=None):
    rclpy.init(args=args)
    try:
        gps_node = GPSNode()
        rclpy.spin(gps_node)
    except KeyboardInterrupt:
        pass
    finally:
        rclpy.shutdown()

if __name__ == '__main__':
    main()
